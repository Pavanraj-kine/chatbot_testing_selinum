from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time

# Config
CHATBOT_URL = 'http://localhost:5000'
EXCEL_FILE = 'questions.xlsx'
INPUT_SELECTOR = '#user-input'
SEND_BUTTON_SELECTOR = '#send-btn'
RESPONSE_SELECTOR = '.chat-box > div'

# Load questions
wb = load_workbook(EXCEL_FILE)
sheet = wb.active

driver = webdriver.Chrome()
wait = WebDriverWait(driver, 600) 

try:
    driver.get(CHATBOT_URL)

    for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        question = row[0]
        expected_response = row[1] if len(row) > 1 else None

        # Wait for input and send button
        input_box = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, INPUT_SELECTOR)))
        send_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, SEND_BUTTON_SELECTOR)))

        # Count current bot responses
        old_responses = driver.find_elements(By.CSS_SELECTOR, RESPONSE_SELECTOR)
        old_count = len(old_responses)

        # Send the question
        input_box.clear()
        input_box.send_keys(question)
        send_button.click()

        # Wait until a new message appears
        def new_response_loaded(driver):
            elements = driver.find_elements(By.CSS_SELECTOR, RESPONSE_SELECTOR)
            return len(elements) > old_count + 1
        # Wait for the new response to load

        wait.until(new_response_loaded)
        new_responses = driver.find_elements(By.CSS_SELECTOR, RESPONSE_SELECTOR)
        last_response = new_responses[-1].text

        print(f"Q: {question}")
        print(f"A: {last_response}")
        if expected_response:
            if expected_response.lower() in last_response.lower():
                print("✅ Matched expected response")
            else:
                print(f"❌ Mismatch. Expected: {expected_response}")
        print("-" * 100)
        with open('chatbot_test_results.txt', 'a') as f:
            f.write(f"Q: {question}\n")
            f.write(f"A: {last_response}\n")
            if expected_response:
                if expected_response.lower() in last_response.lower():
                    f.write("✅ Matched expected response\n")
                else:
                    f.write(f"❌ Mismatch. Expected: {expected_response}\n")
            f.write("-" * 100 + "\n")


finally:
    driver.quit()
